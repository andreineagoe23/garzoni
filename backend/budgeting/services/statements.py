"""
Bank statement (CSV) parsing.

Turns an uploaded statement from any of the banks our users actually have
(Revolut, Barclays, Monzo, Starling, Banca Transilvania, ING RO, BCR, …) into
the internal :class:`budgeting.models.Transaction` shape without needing an
open-banking licence.

Design notes
------------
* **Nothing touches disk.** The uploaded bytes are decoded, parsed and thrown
  away. Only normalised rows are ever persisted.
* **Header mapping is score-based**, not a per-bank hardcode. Known dialects
  only supply hints and a display name — an unknown bank still parses as long
  as it has a date, a description and an amount (or debit/credit pair).
* **Sign convention** matches ``Transaction``: negative = outflow, positive =
  inflow. Statements disagree wildly on this; :func:`_resolve_amounts` is the
  single place that decides.
* **Redaction happens at parse time** so a PAN or IBAN pasted into a memo field
  never reaches the database, the AI layer, or a log line.
"""

from __future__ import annotations

import csv
import io
import logging
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Dict, List, Optional, Sequence, Tuple

# ---------------------------------------------------------------------------
# Limits (also enforced at the view layer before we get here)
# ---------------------------------------------------------------------------

MAX_PREVIEW_ROWS = 5000
SNIFF_BYTES = 8192
HEADER_SCAN_LINES = 15

ENCODINGS = ("utf-8-sig", "utf-8", "cp1252", "iso-8859-2", "latin-1")
DELIMITERS = (",", ";", "\t", "|")


class StatementParseError(Exception):
    """Raised when a file cannot be understood at all."""

    def __init__(self, code: str, message: str) -> None:
        super().__init__(message)
        self.code = code
        self.message = message


# ---------------------------------------------------------------------------
# Redaction
# ---------------------------------------------------------------------------

# 13-19 digits, optionally space/dash grouped: card PANs.
_PAN_RE = re.compile(r"\b(?:\d[ -]?){12,18}\d\b")
# IBAN: 2 letters + 2 digits + up to 30 alphanumerics.
_IBAN_RE = re.compile(r"\b[A-Z]{2}\d{2}[A-Z0-9]{10,30}\b")
# Long digit runs that are not obviously amounts (account numbers, sort codes).
_LONG_DIGITS_RE = re.compile(r"\b\d{9,}\b")


def redact(text: str) -> str:
    """Strip anything that looks like a card number, IBAN or account number.

    Applied to every free-text field before persistence. Deliberately
    aggressive: a mangled merchant string is cheaper than storing a PAN.
    """
    if not text:
        return ""
    out = _IBAN_RE.sub("[account]", text)
    out = _PAN_RE.sub("[card]", out)
    out = _LONG_DIGITS_RE.sub("[number]", out)
    return " ".join(out.split())[:256]


# ---------------------------------------------------------------------------
# Column aliases
# ---------------------------------------------------------------------------

# Each canonical field maps to substrings we look for in a normalised header.
# Earlier entries within a field score higher, and fields are resolved in
# declaration order against a shared pool of headers — so ``debit``/``credit``
# are declared before ``amount``, otherwise the generic "suma"/"paid out"
# aliases would swallow a "Suma debit" column and lose the pairing.
COLUMN_ALIASES: Dict[str, Sequence[str]] = {
    "date": (
        "completed date",
        "transaction date",
        "date of transaction",
        "booking date",
        "value date",
        "data tranzactiei",
        "data inregistrarii",
        "data procesarii",
        "data valuta",
        "posted",
        "date",
        "data",
    ),
    "description": (
        "description",
        "counter party",
        "counterparty",
        "merchant",
        "payee",
        "name",
        "narrative",
        "details",
        "detalii tranzactie",
        "detalii",
        "descriere",
        "explicatii",
        "beneficiar",
        "reference",
        "memo",
    ),
    "debit": ("suma debit", "debit amount", "paid out", "money out", "debit"),
    "credit": ("suma credit", "credit amount", "paid in", "money in", "credit"),
    "amount": (
        "amount (gbp)",
        "amount (eur)",
        "amount (ron)",
        "amount",
        "suma",
        "valoare",
        "value",
    ),
    "currency": ("currency", "valuta", "moneda", "ccy"),
    "balance": ("balance", "sold", "running balance"),
    "category": ("category", "spending category", "categorie", "type of transaction"),
    "kind": ("type", "tip", "dc", "debit/credit", "transaction type"),
    "state": ("state", "status", "stare"),
}

# Fields where a *shorter* alias should not beat a longer, more specific one.
_AMBIGUOUS = {"amount", "date", "data", "name", "type", "reference"}


@dataclass(frozen=True)
class Dialect:
    """A recognised bank export. Only affects labelling and a few quirks."""

    slug: str
    label: str
    # Header fragments that, if all present, identify this bank.
    signature: Sequence[str]
    # Rows to drop, e.g. Revolut's reverted/pending states.
    drop_states: Sequence[str] = ()
    default_currency: str = ""


DIALECTS: Sequence[Dialect] = (
    Dialect(
        slug="revolut",
        label="Revolut",
        signature=("type", "product", "started date", "completed date", "amount"),
        drop_states=("reverted", "declined", "failed"),
    ),
    Dialect(
        slug="monzo",
        label="Monzo",
        signature=("transaction id", "date", "type", "name", "amount"),
    ),
    Dialect(
        slug="starling",
        label="Starling",
        signature=("counter party", "reference", "spending category"),
        default_currency="GBP",
    ),
    Dialect(
        slug="barclays",
        label="Barclays",
        signature=("number", "date", "account", "amount", "memo"),
        default_currency="GBP",
    ),
    Dialect(
        slug="banca-transilvania",
        label="Banca Transilvania",
        signature=("data", "descriere"),
        default_currency="RON",
    ),
    Dialect(
        slug="ing-ro",
        label="ING Romania",
        signature=("data", "detalii tranzactie"),
        default_currency="RON",
    ),
)


# ---------------------------------------------------------------------------
# Parsed output
# ---------------------------------------------------------------------------


@dataclass
class ParsedRow:
    posted_at: date
    amount: Decimal
    currency: str
    description: str
    merchant_name: str = ""
    raw_category: str = ""
    fingerprint: str = ""


@dataclass
class ParsedStatement:
    rows: List[ParsedRow]
    dialect_slug: str
    dialect_label: str
    currency: str
    column_map: Dict[str, str]
    total_rows_seen: int = 0
    skipped_rows: int = 0
    warnings: List[str] = field(default_factory=list)
    #: csv | xlsx | pdf | ofx | qif — surfaced so the UI can say what it read.
    source_format: str = "csv"

    @property
    def period_start(self) -> Optional[date]:
        return min((r.posted_at for r in self.rows), default=None)

    @property
    def period_end(self) -> Optional[date]:
        return max((r.posted_at for r in self.rows), default=None)


# ---------------------------------------------------------------------------
# Decoding & dialect sniffing
# ---------------------------------------------------------------------------


def detect_format(raw: bytes) -> str:
    """Sniff the container format from magic bytes and leading content.

    Extension is not trusted: banks mislabel downloads, and mobile pickers
    hand us ``application/octet-stream`` for everything.
    """
    head = raw[:512]
    if head[:5] == b"%PDF-":
        return "pdf"
    if head[:4] == b"PK\x03\x04":
        # xlsx is a zip; .xls (OLE2) is not, and we do not support it.
        return "xlsx"
    if head[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        raise StatementParseError(
            "legacy_excel",
            "That's an old .xls file. Re-save it as .xlsx or CSV and try again.",
        )
    probe = head.lstrip().upper()
    if probe.startswith(b"OFXHEADER") or b"<OFX>" in raw[:4096].upper():
        return "ofx"
    if probe.startswith(b"!TYPE:"):
        return "qif"
    return "csv"


def decode(raw: bytes) -> str:
    """Decode statement bytes, trying the encodings European banks actually use."""
    for enc in ENCODINGS:
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    raise StatementParseError("undecodable", "Could not read that file as text.")


def _sniff_delimiter(sample: str) -> str:
    try:
        return csv.Sniffer().sniff(sample[:SNIFF_BYTES], delimiters="".join(DELIMITERS)).delimiter
    except csv.Error:
        # Sniffer gives up on short or ragged files — fall back to whichever
        # candidate appears most often in the first few lines.
        head = "\n".join(sample.splitlines()[:HEADER_SCAN_LINES])
        counts = {d: head.count(d) for d in DELIMITERS}
        best = max(counts, key=lambda d: counts[d])
        return best if counts[best] else ","


def _normalise_header(value: str) -> str:
    return " ".join(str(value or "").strip().lower().replace("_", " ").split())


def _find_header_row(rows: Sequence[Sequence[str]]) -> int:
    """Return the index of the most header-like row in the first N lines.

    Barclays and several Romanian exports prepend account metadata before the
    real header, so we cannot assume row 0.
    """
    best_idx, best_score = 0, -1
    for idx, row in enumerate(rows[:HEADER_SCAN_LINES]):
        cells = [_normalise_header(c) for c in row if str(c).strip()]
        if len(cells) < 2:
            continue
        score = 0
        for field_name, aliases in COLUMN_ALIASES.items():
            if any(any(a in cell for a in aliases) for cell in cells):
                score += 2 if field_name in ("date", "description", "amount") else 1
        # Headers are text, not numbers.
        numeric = sum(1 for c in cells if re.fullmatch(r"[-+\d.,\s]+", c))
        score -= numeric
        if score > best_score:
            best_idx, best_score = idx, score
    if best_score <= 0:
        raise StatementParseError(
            "no_header",
            "Couldn't find a header row. Make sure the CSV includes column names.",
        )
    return best_idx


def _map_columns(header: Sequence[str]) -> Dict[str, str]:
    """Map canonical field -> actual header name, best match wins."""
    normalised = [(_normalise_header(h), h) for h in header]
    mapping: Dict[str, str] = {}
    taken: set = set()
    for field_name, aliases in COLUMN_ALIASES.items():
        best: Optional[Tuple[int, str]] = None
        for norm, original in normalised:
            if not norm or original in taken:
                continue
            for rank, alias in enumerate(aliases):
                if alias == norm:
                    score = 1000 - rank
                elif alias in norm:
                    # A generic alias inside a longer header is a weak signal.
                    score = (500 - rank) - (100 if alias in _AMBIGUOUS else 0)
                else:
                    continue
                if best is None or score > best[0]:
                    best = (score, original)
                break
        if best is not None:
            mapping[field_name] = best[1]
            taken.add(best[1])
    return mapping


def _detect_dialect(header: Sequence[str]) -> Optional[Dialect]:
    cells = {_normalise_header(h) for h in header}
    for dialect in DIALECTS:
        if all(any(sig in cell for cell in cells) for sig in dialect.signature):
            return dialect
    return None


# ---------------------------------------------------------------------------
# Value coercion
# ---------------------------------------------------------------------------

_DATE_FORMATS = (
    "%Y-%m-%d",
    "%d/%m/%Y",
    "%d.%m.%Y",
    "%d-%m-%Y",
    "%m/%d/%Y",
    "%Y/%m/%d",
    "%d %b %Y",
    "%d %B %Y",
    "%b %d, %Y",
    # Two-digit years: the printed form on most UK PDF statements ("02 Jul 26")
    # and on some short-form CSV exports.
    "%d/%m/%y",
    "%d.%m.%y",
    "%d-%m-%y",
    "%d %b %y",
    "%d %B %y",
)


def parse_date(value: str) -> Optional[date]:
    """Parse a statement date, tolerating an appended time component.

    Ambiguous ``dd/mm`` vs ``mm/dd`` is resolved day-first (UK + RO exports);
    the US ordering is only tried when day-first is impossible.
    """
    text = str(value or "").strip()
    if not text:
        return None

    # Try the whole string first: spelled-out dates ("02 Jul 26") contain
    # spaces, so splitting on whitespace up front would destroy them. Only
    # fall back to the leading token, which strips an ISO time component.
    candidates = [text]
    head = text.replace("T", " ").split(" ")[0].strip()
    if head and head != text:
        candidates.append(head)

    for candidate in candidates:
        for fmt in _DATE_FORMATS:
            try:
                return datetime.strptime(candidate, fmt).date()
            except ValueError:
                continue
    try:  # ISO with offset, e.g. 2026-07-03T10:15:00+01:00
        return datetime.fromisoformat(text).date()
    except (ValueError, TypeError):
        return None


_AMOUNT_CLEAN_RE = re.compile(r"[^\d,.\-+()]")


def parse_amount(value: str) -> Optional[Decimal]:
    """Parse UK (``1,234.56``) and continental (``1.234,56``) amounts.

    Handles currency symbols, thin spaces, trailing ``CR``/``DB`` markers and
    accountant parentheses for negatives.
    """
    text = str(value or "").strip()
    if not text:
        return None
    lowered = text.lower()
    trailing_credit = lowered.endswith(("cr", "credit"))
    trailing_debit = lowered.endswith(("db", "dr", "debit"))
    cleaned = _AMOUNT_CLEAN_RE.sub("", text)
    if not cleaned:
        return None
    negative = cleaned.startswith("-") or ("(" in cleaned and ")" in cleaned)
    cleaned = cleaned.replace("(", "").replace(")", "").lstrip("+-")
    if not cleaned:
        return None

    has_comma = "," in cleaned
    has_dot = "." in cleaned
    if has_comma and has_dot:
        # Whichever separator comes last is the decimal one.
        if cleaned.rfind(",") > cleaned.rfind("."):
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    elif has_comma or has_dot:
        sep = "," if has_comma else "."
        head, _, tail = cleaned.rpartition(sep)
        # Statement amounts carry 0 or 2 decimals, so exactly 3 trailing digits
        # means thousands grouping ("1.234" and "1,234" are both 1234).
        if len(tail) == 3 and tail.isdigit():
            cleaned = head.replace(sep, "") + tail
        else:
            cleaned = head.replace(sep, "") + "." + tail
    try:
        amount = Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return None
    if negative or trailing_debit:
        amount = -abs(amount)
    elif trailing_credit:
        amount = abs(amount)
    return amount


_CURRENCY_SYMBOLS = {"£": "GBP", "€": "EUR", "$": "USD", "lei": "RON", "ron": "RON"}


def _currency_from(value: str, header: str = "") -> str:
    text = str(value or "").strip().upper()
    if len(text) == 3 and text.isalpha():
        return text
    for symbol, code in _CURRENCY_SYMBOLS.items():
        if symbol.upper() in text or symbol.upper() in header.upper():
            return code
    match = re.search(r"\(([A-Z]{3})\)", header or "")
    return match.group(1) if match else ""


_DEBIT_MARKERS = {"debit", "db", "dr", "d", "out", "withdrawal", "plata", "iesire"}
_CREDIT_MARKERS = {"credit", "cr", "c", "in", "deposit", "incasare", "intrare"}


def _resolve_amounts(
    raw: Dict[str, str],
    column_map: Dict[str, str],
) -> Optional[Decimal]:
    """Produce a signed amount for one row (negative = money out)."""
    debit_col = column_map.get("debit")
    credit_col = column_map.get("credit")
    amount_col = column_map.get("amount")

    # Separate debit/credit columns (Banca Transilvania, Barclays business).
    # Only trust this shape when the two columns are genuinely different —
    # "paid out"/"paid in" can both alias onto a single "amount" header.
    if debit_col and credit_col and debit_col != credit_col:
        debit = parse_amount(raw.get(debit_col, ""))
        credit = parse_amount(raw.get(credit_col, ""))
        if debit:
            return -abs(debit)
        if credit:
            return abs(credit)
        return None

    if not amount_col:
        return None
    amount = parse_amount(raw.get(amount_col, ""))
    if amount is None:
        return None

    # Unsigned amount plus a separate debit/credit marker column.
    kind_col = column_map.get("kind")
    if kind_col and amount > 0:
        marker = _normalise_header(raw.get(kind_col, ""))
        if marker in _DEBIT_MARKERS:
            return -amount
        if marker in _CREDIT_MARKERS:
            return amount
    return amount


def _fingerprint(row: ParsedRow) -> str:
    """Stable per-row identity used for deduplication across re-uploads.

    Deliberately excludes the source file, so re-importing an overlapping
    statement updates rather than duplicates.
    """
    import hashlib

    seed = f"{row.posted_at.isoformat()}|{row.amount}|{row.currency}|{row.description.lower()}"
    return "csv_" + hashlib.sha256(seed.encode("utf-8")).hexdigest()[:40]


# ---------------------------------------------------------------------------
# Shared finalisation
# ---------------------------------------------------------------------------


def _finalize(
    rows: List[ParsedRow],
    *,
    dialect_slug: str,
    dialect_label: str,
    source_format: str,
    column_map: Optional[Dict[str, str]] = None,
    seen: int = 0,
    skipped: int = 0,
    warnings: Optional[List[str]] = None,
    truncated: bool = False,
    max_rows: int = MAX_PREVIEW_ROWS,
    allow_sign_flip: bool = True,
) -> ParsedStatement:
    """Common tail for every format: warnings, sign sanity, currency, fingerprints."""
    warnings = list(warnings or [])
    if not rows:
        raise StatementParseError(
            "no_transactions",
            "No transactions could be read from that file. Check it's a statement export.",
        )

    if truncated:
        warnings.append(f"Only the first {max_rows} transactions were read.")
    if skipped:
        warnings.append(f"{skipped} row(s) were skipped because they had no usable date or amount.")

    # A statement where every amount is positive almost always means an
    # unsigned "spending" export. Flip it, and say so. Formats that carry an
    # explicit direction per transaction (OFX, QIF, debit/credit columns) opt
    # out via allow_sign_flip.
    if allow_sign_flip and all(r.amount > 0 for r in rows):
        for row in rows:
            row.amount = -row.amount
        warnings.append(
            "All amounts were positive, so they were read as money out. "
            "Check the totals below before saving."
        )

    currencies = {r.currency for r in rows if r.currency}
    dominant = ""
    if currencies:
        dominant = max(currencies, key=lambda c: sum(1 for r in rows if r.currency == c))
        if len(currencies) > 1:
            warnings.append(
                f"This statement mixes {len(currencies)} currencies; totals are shown in {dominant}."
            )
    for row in rows:
        if not row.currency:
            row.currency = dominant
        row.fingerprint = _fingerprint(row)

    return ParsedStatement(
        rows=rows,
        dialect_slug=dialect_slug,
        dialect_label=dialect_label,
        currency=dominant,
        column_map=column_map or {},
        total_rows_seen=seen or len(rows),
        skipped_rows=skipped,
        warnings=warnings,
        source_format=source_format,
    )


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------


def parse_statement(raw: bytes, max_rows: int = MAX_PREVIEW_ROWS) -> ParsedStatement:
    """Parse statement bytes into normalised rows.

    Dispatches on the *sniffed* container format, not the file extension.
    Raises :class:`StatementParseError` when the file cannot be understood at
    all; individual bad rows are skipped and counted instead.
    """
    if not raw or not raw.strip():
        raise StatementParseError("empty_file", "That file is empty.")

    fmt = detect_format(raw)
    if fmt == "pdf":
        return parse_pdf(raw, max_rows=max_rows)
    if fmt == "xlsx":
        return _parse_tabular(_rows_from_xlsx(raw), max_rows=max_rows, source_format="xlsx")
    if fmt == "ofx":
        return parse_ofx(decode(raw), max_rows=max_rows)
    if fmt == "qif":
        return parse_qif(decode(raw), max_rows=max_rows)
    return _parse_tabular(_rows_from_csv(decode(raw)), max_rows=max_rows, source_format="csv")


def _rows_from_csv(text: str) -> List[List[str]]:
    if not text.strip():
        raise StatementParseError("empty_file", "That file is empty.")
    delimiter = _sniff_delimiter(text)
    rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
    rows = [r for r in rows if any(str(c).strip() for c in r)]
    if not rows:
        raise StatementParseError("empty_file", "That file has no rows.")
    return rows


def _rows_from_xlsx(raw: bytes) -> List[List[str]]:
    """Read the first worksheet of an .xlsx export into CSV-shaped rows.

    Dates come back as ``datetime`` and amounts as ``float``; both are
    stringified here so the same column mapper and value parsers handle them.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover - dependency is pinned in requirements
        raise StatementParseError(
            "xlsx_unsupported",
            "Spreadsheet statements aren't supported on this server yet.",
        )

    try:
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception:
        raise StatementParseError(
            "xlsx_unreadable",
            "That spreadsheet couldn't be opened. Try exporting as CSV instead.",
        )
    try:
        sheet = workbook[workbook.sheetnames[0]]
        rows: List[List[str]] = []
        for excel_row in sheet.iter_rows(values_only=True):
            cells = []
            for value in excel_row:
                if value is None:
                    cells.append("")
                elif isinstance(value, datetime):
                    cells.append(value.date().isoformat())
                elif isinstance(value, date):
                    cells.append(value.isoformat())
                else:
                    cells.append(str(value))
            if any(c.strip() for c in cells):
                rows.append(cells)
    finally:
        workbook.close()

    if not rows:
        raise StatementParseError("empty_file", "That spreadsheet has no rows.")
    return rows


def _parse_tabular(
    all_rows: List[List[str]],
    *,
    max_rows: int = MAX_PREVIEW_ROWS,
    source_format: str = "csv",
) -> ParsedStatement:
    """Header-mapped parsing shared by CSV and spreadsheet exports."""
    header_idx = _find_header_row(all_rows)
    header = [str(h).strip() for h in all_rows[header_idx]]
    column_map = _map_columns(header)
    dialect = _detect_dialect(header)

    if "date" not in column_map:
        raise StatementParseError("no_date_column", "Couldn't find a date column.")
    if "amount" not in column_map and not ("debit" in column_map and "credit" in column_map):
        raise StatementParseError("no_amount_column", "Couldn't find an amount column.")

    warnings: List[str] = []
    rows: List[ParsedRow] = []
    skipped = 0
    seen = 0
    truncated = False

    default_currency = dialect.default_currency if dialect else ""
    amount_header = column_map.get("amount", "")
    header_currency = _currency_from("", amount_header) or default_currency

    for data_row in all_rows[header_idx + 1 :]:
        seen += 1
        if len(rows) >= max_rows:
            truncated = True
            break
        raw_map = {
            header[i]: (data_row[i] if i < len(data_row) else "")
            for i in range(min(len(header), len(data_row)))
        }
        if not raw_map:
            skipped += 1
            continue

        if dialect and dialect.drop_states and column_map.get("state"):
            state = _normalise_header(raw_map.get(column_map["state"], ""))
            if state in dialect.drop_states:
                skipped += 1
                continue

        posted_at = parse_date(raw_map.get(column_map["date"], ""))
        amount = _resolve_amounts(raw_map, column_map)
        if posted_at is None or amount is None or amount == 0:
            skipped += 1
            continue

        description = redact(raw_map.get(column_map.get("description", ""), ""))
        if not description:
            description = redact(raw_map.get(column_map.get("category", ""), "")) or "Transaction"

        currency = (
            _currency_from(raw_map.get(column_map.get("currency", ""), "")) or header_currency or ""
        )

        rows.append(
            ParsedRow(
                posted_at=posted_at,
                amount=amount,
                currency=currency,
                description=description,
                raw_category=redact(raw_map.get(column_map.get("category", ""), ""))[:128],
            )
        )

    generic_label = "Generic spreadsheet" if source_format == "xlsx" else "Generic CSV"
    return _finalize(
        rows,
        dialect_slug=dialect.slug if dialect else "generic",
        dialect_label=dialect.label if dialect else generic_label,
        source_format=source_format,
        column_map=column_map,
        seen=seen,
        skipped=skipped,
        warnings=warnings,
        truncated=truncated,
        max_rows=max_rows,
        # A debit/credit pair already encodes direction, so an all-positive
        # result there is real, not an unsigned export.
        allow_sign_flip="credit" not in column_map,
    )


# ---------------------------------------------------------------------------
# PDF statements
# ---------------------------------------------------------------------------
#
# Barclays (and most UK high-street banks) only offer PDF for anything older
# than the last few months, so this is the format users actually have.
#
# PDFs carry no column semantics — only glyphs at coordinates. The approach:
#
#   1. Group extracted words into visual lines by their y position.
#   2. A transaction line starts with a date. Continuation lines (wrapped
#      descriptions) do not, and get appended to the previous transaction.
#   3. Numeric tokens are clustered by x position across the whole document
#      into columns. The rightmost recurring column is the running balance.
#   4. Direction comes from the balance delta where a balance exists — the
#      single most reliable signal in a PDF, since "money out" vs "money in"
#      is otherwise only implied by which column a number sits in.

_PDF_DATE_RE = re.compile(
    r"^\s*("
    r"\d{1,2}[/.\-]\d{1,2}[/.\-]\d{2,4}"
    r"|\d{1,2}\s+[A-Za-z]{3,9}\s*\d{0,4}"
    r"|\d{4}-\d{2}-\d{2}"
    r")\b"
)
# Money tokens must carry 2 decimals. That is exactly what separates an amount
# from the day and year in "02 Jul 26", which would otherwise be read as two
# extra numeric columns on every single transaction line.
_PDF_NUMBER_RE = re.compile(
    r"^[\-+(]?[£€$]?\s?"
    r"(?:\d{1,3}(?:[.,\s]\d{3})+|\d+)"  # grouped thousands, or plain digits
    r"[.,]\d{2}"
    r"\)?(?:CR|DR|DB)?$",
    re.IGNORECASE,
)
# Rows on a bank PDF are ~8-12pt tall; anything within this counts as one line.
_PDF_LINE_TOLERANCE = 3.0
_PDF_COLUMN_TOLERANCE = 12.0
MAX_PDF_PAGES = 40
# A wrapped description is at most a line or two below its transaction.
_PDF_CONTINUATION_GAP = _PDF_LINE_TOLERANCE * 5
_PDF_CONTINUATION_MAX_CHARS = 60

# Real Barclays statements print the transaction date as "15 Jul" with no year
# — the year only appears once, in the statement header. Without recovering it
# every single row would fail to parse.
_PDF_DAY_MONTH_RE = re.compile(r"^\s*(\d{1,2})\s+([A-Za-z]{3,9})\s*$")

# Statement furniture that sits on a dated line but is not a transaction:
# opening/closing balances and page carry-overs. Importing these as spending
# silently corrupts every total.
_PDF_BALANCE_MARKER_RE = re.compile(
    r"\b("
    r"start(ing)?\s+balance|end(ing)?\s+balance|opening\s+balance|closing\s+balance"
    r"|balance\s+(brought|carried)\s+forward|brought\s+forward|carried\s+forward"
    r"|b/?fwd|c/?fwd|total\s+(payments|receipts)"
    r")\b",
    re.IGNORECASE,
)
_YEAR_RE = re.compile(r"\b(19|20)\d{2}\b")


def _infer_statement_year(lines: Sequence[_PdfLine]) -> Optional[int]:
    """Most frequent plausible year printed anywhere in the document."""
    counts: Dict[int, int] = {}
    for line in lines:
        for match in _YEAR_RE.finditer(line.text):
            year = int(match.group(0))
            if 1990 <= year <= date.today().year + 1:
                counts[year] = counts.get(year, 0) + 1
    if not counts:
        return None
    # On a tie, take the *earliest* year: rows run forward through the
    # statement, so a period spanning new year should start in the older one
    # and roll forward (handled at the row level), never start in the newer.
    return max(counts.items(), key=lambda kv: (kv[1], -kv[0]))[0]


def _pdf_date(token: str, fallback_year: Optional[int]) -> Optional[date]:
    """Parse a PDF date token, filling in a missing year from the header."""
    parsed = parse_date(token)
    if parsed is not None:
        return parsed
    match = _PDF_DAY_MONTH_RE.match(token)
    if not match or fallback_year is None:
        return None
    return parse_date(f"{match.group(1)} {match.group(2)} {fallback_year}")


@dataclass
class _PdfLine:
    top: float
    text: str
    numbers: List[Tuple[Decimal, float]]  # (value, x centre)
    page: int = 0


def _pdf_lines(page, page_number: int = 0) -> List[_PdfLine]:
    """Group a page's words into visual lines, keeping numeric x positions."""
    words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
    buckets: Dict[int, List[dict]] = {}
    for word in words:
        key = int(round(float(word["top"]) / _PDF_LINE_TOLERANCE))
        buckets.setdefault(key, []).append(word)

    lines: List[_PdfLine] = []
    for key in sorted(buckets):
        group = sorted(buckets[key], key=lambda w: float(w["x0"]))
        text = " ".join(w["text"] for w in group)
        numbers: List[Tuple[Decimal, float]] = []
        for word in group:
            token = word["text"].strip()
            if not _PDF_NUMBER_RE.match(token):
                continue
            value = parse_amount(token)
            if value is None:
                continue
            centre = (float(word["x0"]) + float(word["x1"])) / 2
            numbers.append((value, centre))
        lines.append(
            _PdfLine(
                top=key * _PDF_LINE_TOLERANCE,
                text=text,
                numbers=numbers,
                page=page_number,
            )
        )
    return lines


def _cluster_columns(centres: Sequence[float]) -> List[float]:
    """Collapse x centres into column positions (simple 1-D agglomeration)."""
    columns: List[float] = []
    for centre in sorted(centres):
        if columns and centre - columns[-1] <= _PDF_COLUMN_TOLERANCE:
            columns[-1] = (columns[-1] + centre) / 2
        else:
            columns.append(centre)
    return columns


def _pdf_balance_column(lines: Sequence[_PdfLine]) -> Optional[float]:
    """Identify the running-balance column: the rightmost column present on
    most transaction lines. Returns ``None`` when there clearly isn't one."""
    tx_lines = [ln for ln in lines if _PDF_DATE_RE.match(ln.text) and ln.numbers]
    if len(tx_lines) < 3:
        return None
    columns = _cluster_columns([c for ln in tx_lines for _, c in ln.numbers])
    if len(columns) < 2:
        return None
    rightmost = columns[-1]
    hits = sum(
        1
        for ln in tx_lines
        if any(abs(c - rightmost) <= _PDF_COLUMN_TOLERANCE for _, c in ln.numbers)
    )
    return rightmost if hits >= len(tx_lines) * 0.8 else None


def _pdf_money_in_column(lines: Sequence[_PdfLine], balance_x: Optional[float]) -> Optional[float]:
    """Locate a separate "money in" / "paid in" column, if the layout has one.

    Barclays and Lloyds print money out and money in as two columns to the left
    of the balance. When that layout is present, which column a number sits in
    *is* the direction — far more reliable than inferring it. Returns the x
    centre of the money-in (rightmost non-balance) column, or ``None`` when the
    statement uses a single amount column.
    """
    centres: List[float] = []
    for line in lines:
        if not _PDF_DATE_RE.match(line.text):
            continue
        for _, centre in line.numbers:
            if balance_x is not None and abs(centre - balance_x) <= _PDF_COLUMN_TOLERANCE:
                continue
            centres.append(centre)

    columns = _cluster_columns(centres)
    if len(columns) < 2:
        return None

    # Guard against a reference-number column masquerading as money: the two
    # money columns sit close together, well right of the description.
    if columns[-1] - columns[-2] > _PDF_COLUMN_TOLERANCE * 12:
        return None
    return columns[-1]


def parse_pdf(raw: bytes, max_rows: int = MAX_PREVIEW_ROWS) -> ParsedStatement:
    """Extract transactions from a text-based PDF statement."""
    try:
        import pdfplumber
    except ImportError:  # pragma: no cover - dependency is pinned in requirements
        raise StatementParseError(
            "pdf_unsupported",
            "PDF statements aren't supported on this server yet.",
        )

    # pdfminer emits a DEBUG record per token — tens of thousands of lines for
    # one statement, and the logging call itself dominates the parse time.
    # settings.LOGGING pins this too; this guards runners with their own config.
    for noisy in ("pdfminer", "pdfplumber"):
        logging.getLogger(noisy).setLevel(logging.WARNING)

    lines: List[_PdfLine] = []
    try:
        with pdfplumber.open(io.BytesIO(raw)) as pdf:
            if len(pdf.pages) > MAX_PDF_PAGES:
                raise StatementParseError(
                    "pdf_too_long",
                    f"That PDF has more than {MAX_PDF_PAGES} pages. "
                    "Export a shorter date range.",
                )
            for page_number, page in enumerate(pdf.pages):
                lines.extend(_pdf_lines(page, page_number))
    except StatementParseError:
        raise
    except Exception as exc:
        message = str(exc).lower()
        if "password" in message or "encrypt" in message:
            raise StatementParseError(
                "pdf_encrypted",
                "That PDF is password protected. Remove the password and try again.",
            )
        raise StatementParseError(
            "pdf_unreadable",
            "That PDF couldn't be read. If your bank also offers CSV, use that instead.",
        )

    if not any(ln.text.strip() for ln in lines):
        raise StatementParseError(
            "pdf_no_text",
            "That PDF has no selectable text — it looks like a scan. "
            "Download the statement again from your bank, or use a CSV export.",
        )

    balance_x = _pdf_balance_column(lines)
    money_in_x = _pdf_money_in_column(lines, balance_x)
    fallback_year = _infer_statement_year(lines)
    warnings: List[str] = []
    rows: List[ParsedRow] = []
    seen = 0
    skipped = 0
    truncated = False
    previous_balance: Optional[Decimal] = None
    last_row_anchor: Optional[Tuple[int, float]] = None
    directed = 0
    assumed = 0

    for line in lines:
        match = _PDF_DATE_RE.match(line.text)
        if not match:
            # A wrapped description sits directly under its transaction, on the
            # same page. Anything further away is page furniture — footers,
            # legal blurb, "Continued" markers — and must not be glued on.
            if (
                rows
                and last_row_anchor is not None
                and line.text.strip()
                and not line.numbers
                and line.page == last_row_anchor[0]
                and 0 < line.top - last_row_anchor[1] <= _PDF_CONTINUATION_GAP
                and len(line.text.strip()) <= _PDF_CONTINUATION_MAX_CHARS
                and not _PDF_BALANCE_MARKER_RE.search(line.text)
            ):
                extra = redact(line.text.strip())
                if extra:
                    rows[-1].description = f"{rows[-1].description} {extra}"[:256]
                    last_row_anchor = (line.page, line.top)
            continue

        seen += 1
        if len(rows) >= max_rows:
            truncated = True
            break

        if _PDF_BALANCE_MARKER_RE.search(line.text):
            # Keep the balance itself so the delta chain stays correct, but
            # never emit it as a transaction.
            if balance_x is not None:
                for value, centre in line.numbers:
                    if abs(centre - balance_x) <= _PDF_COLUMN_TOLERANCE:
                        previous_balance = value
                        break
            skipped += 1
            continue

        posted_at = _pdf_date(match.group(1), fallback_year)
        if posted_at is None or not line.numbers:
            skipped += 1
            continue

        # A statement that runs across new year prints "28 Dec" then "03 Jan";
        # with a single inferred year the second one lands 12 months early.
        if rows and (rows[-1].posted_at - posted_at).days > 180:
            posted_at = posted_at.replace(year=posted_at.year + 1)

        balance: Optional[Decimal] = None
        amounts = list(line.numbers)
        if balance_x is not None:
            for index, (value, centre) in enumerate(amounts):
                if abs(centre - balance_x) <= _PDF_COLUMN_TOLERANCE:
                    balance = value
                    amounts.pop(index)
                    break
        if not amounts:
            skipped += 1
            continue

        # The amount is the number nearest the balance column (the last money
        # column before it); anything further left is a reference number.
        amount, amount_x = amounts[-1]

        # Direction, best signal first:
        #   1. A separate "money in" column — Barclays and Lloyds print one.
        #   2. The running-balance delta.
        #   3. Nothing left: assume money out, and say how often we had to.
        if money_in_x is not None and abs(amount_x - money_in_x) <= _PDF_COLUMN_TOLERANCE:
            amount = abs(amount)
            directed += 1
        elif money_in_x is not None:
            amount = -abs(amount)
            directed += 1
        elif balance is not None and previous_balance is not None and balance != previous_balance:
            amount = -abs(amount) if balance < previous_balance else abs(amount)
            directed += 1
        else:
            amount = -abs(amount)
            assumed += 1

        if balance is not None:
            previous_balance = balance

        description = redact(
            line.text[match.end() :].strip()
            # Strip the trailing numeric run so the description is words only.
            .rsplit(" ", len(line.numbers))[0]
        )
        if not description:
            description = "Transaction"

        rows.append(
            ParsedRow(
                posted_at=posted_at,
                amount=amount,
                currency="",
                description=description,
            )
        )
        last_row_anchor = (line.page, line.top)

    if assumed:
        warnings.append(
            f"{assumed} line(s) had no money-in column or balance to check against "
            "and were read as money out. Check the sample below."
        )
    warnings.append(
        "PDF statements are read from the printed layout, so they are less "
        "exact than a CSV export. If your bank also offers CSV or OFX, prefer that."
    )

    return _finalize(
        rows,
        dialect_slug="pdf",
        dialect_label="PDF statement",
        source_format="pdf",
        seen=seen,
        skipped=skipped,
        warnings=warnings,
        truncated=truncated,
        max_rows=max_rows,
        # Direction is already resolved per line above; a blanket flip would
        # undo it on any statement that happens to be all inflows.
        allow_sign_flip=False,
    )


# ---------------------------------------------------------------------------
# OFX / QIF
# ---------------------------------------------------------------------------
#
# Barclays, Lloyds, HSBC and Nationwide all offer "Quicken/Money" downloads.
# These are far more reliable than PDF — every field is explicit — so when a
# user has the choice this is the best thing they can upload.

_OFX_TAG_RE = re.compile(r"<([A-Z0-9.]+)>([^<\r\n]*)", re.IGNORECASE)


def parse_ofx(text: str, max_rows: int = MAX_PREVIEW_ROWS) -> ParsedStatement:
    """Parse OFX/QFX (SGML or XML flavour) statement downloads."""
    currency = ""
    currency_match = re.search(r"<CURDEF>\s*([A-Z]{3})", text, re.IGNORECASE)
    if currency_match:
        currency = currency_match.group(1).upper()

    blocks = re.split(r"<STMTTRN>", text, flags=re.IGNORECASE)[1:]
    rows: List[ParsedRow] = []
    skipped = 0
    truncated = False

    for block in blocks:
        if len(rows) >= max_rows:
            truncated = True
            break
        block = re.split(r"</STMTTRN>", block, flags=re.IGNORECASE)[0]
        fields: Dict[str, str] = {}
        for tag, value in _OFX_TAG_RE.findall(block):
            key = tag.upper()
            if key not in fields:
                fields[key] = value.strip()

        posted_raw = fields.get("DTPOSTED", "")[:8]
        try:
            posted_at = datetime.strptime(posted_raw, "%Y%m%d").date()
        except ValueError:
            skipped += 1
            continue

        amount = parse_amount(fields.get("TRNAMT", ""))
        if amount is None or amount == 0:
            skipped += 1
            continue

        description = redact(fields.get("NAME") or fields.get("MEMO") or "Transaction")
        rows.append(
            ParsedRow(
                posted_at=posted_at,
                amount=amount,
                currency=currency,
                description=description or "Transaction",
                raw_category=redact(fields.get("TRNTYPE", ""))[:128],
            )
        )

    return _finalize(
        rows,
        dialect_slug="ofx",
        dialect_label="OFX/QFX download",
        source_format="ofx",
        skipped=skipped,
        truncated=truncated,
        max_rows=max_rows,
        # TRNAMT is signed by the spec.
        allow_sign_flip=False,
    )


def parse_qif(text: str, max_rows: int = MAX_PREVIEW_ROWS) -> ParsedStatement:
    """Parse QIF (Quicken Interchange Format) downloads."""
    rows: List[ParsedRow] = []
    skipped = 0
    truncated = False
    current: Dict[str, str] = {}

    def flush() -> None:
        nonlocal skipped
        if not current:
            return
        posted_at = parse_date(current.get("D", "").replace("'", "/"))
        amount = parse_amount(current.get("T") or current.get("U") or "")
        if posted_at is None or amount is None or amount == 0:
            skipped += 1
            return
        description = redact(current.get("P") or current.get("M") or "Transaction")
        rows.append(
            ParsedRow(
                posted_at=posted_at,
                amount=amount,
                currency="",
                description=description or "Transaction",
                raw_category=redact(current.get("L", ""))[:128],
            )
        )

    for line in text.splitlines():
        line = line.strip()
        if not line or line.startswith("!"):
            continue
        if line == "^":
            if len(rows) >= max_rows:
                truncated = True
                break
            flush()
            current = {}
            continue
        code, value = line[0].upper(), line[1:].strip()
        current.setdefault(code, value)
    else:
        if len(rows) < max_rows:
            flush()

    return _finalize(
        rows,
        dialect_slug="qif",
        dialect_label="QIF download",
        source_format="qif",
        skipped=skipped,
        truncated=truncated,
        max_rows=max_rows,
        # QIF T amounts are signed.
        allow_sign_flip=False,
    )
